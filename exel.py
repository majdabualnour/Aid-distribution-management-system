import openpyxl 


# Give the location of the file 
path = "jj.xlsx"

wb_obj = openpyxl.load_workbook(path) 

sheet_obj = wb_obj.active 

row = sheet_obj.max_row 
column = sheet_obj.max_column 

# print("Total Rows:", row) 
# print("Total Columns:", column) 
# def fdf():
# 	print("\nValue of first column") 
# 	for i in range(1, row + 1): 
# 		cell_obj = sheet_obj.cell(row=i, column=1) 
# 		print(cell_obj.value) 

# def get_dat():
# 	temp= []
# 	for i in range(1, column + 1): 
# 		cell_obj = sheet_obj.cell(row=1, column=i) 
# 		temp.append(cell_obj.value)
# 	return temp
def get_datfi():
	temp= []
	for i in range(1, column + 1): 
		cell_obj = sheet_obj.cell(row=1, column=i) 
		temp.append(cell_obj.value)
	return temp
def applyto(maj):
    maj = maj.replace('أ' , 'ا')
    maj = maj.replace('إ' , 'ا')
    maj = maj.replace('ة','ه')
    return maj
def get_dat():
	temp= []
	tempp = []
	for dd in range(2, 20): 
		for i in range(1, column + 1): 

			cell_obj = sheet_obj.cell(row=dd, column=i) 
			temp.append(cell_obj.value)
		# tempp.append(temp)
		if temp != []:
			cell_obj = sheet_obj.cell(row=dd, column=28)
			meme = cell_obj.value
			if meme == 'استلم':
				
				temp.append('y')
			else:
				temp.append('n')
			
			
			tempp.append(temp)
			temp= []
		temp= []
	return tempp
def ssds(dd, ss):
	dd = str(dd)

	temp= []
	tempp = []
	val = 0
	vall = []
	ss= str(ss)
	if ss == 'all':
		for d in range(2, row + 1): 
			for i in range(1, column + 2): 
				
				cell_obj = sheet_obj.cell(row=d, column=i) 
				
				tempfortem =applyto(str(cell_obj.value))
				
				if dd in tempfortem : 
					val+=1

				temp.append(cell_obj.value)
			if val== 0 :
				temp = []
			val = 0
			
			if temp != []:
				cell_obj = sheet_obj.cell(row=d, column=28)
				meme = cell_obj.value
				if meme == 'استلم':
					
					temp.append('y')
				else:
					temp.append('n')
				
                
				tempp.append(temp)
				temp= []

	

		
	else:
		for i in range(1, column + 1): 
			if str(sheet_obj.cell(row=1, column=i).value) == str(ss):
				
				for d in range(2, row +1): 
					cell_obj = sheet_obj.cell(row=d, column=i)
				 
					
					if dd in str(cell_obj.value) : 
						vall.append(d)
		print(vall)
		for i in vall :

			for sf in range(1,column+1):
				cell_obj = sheet_obj.cell(row=i, column=sf)
				temp.append(str(cell_obj.value))
			
			
			if temp != []:
				
				cell_obj = sheet_obj.cell(row=i, column=28)
				meme = cell_obj.value
				if meme == 'استلم':
					
					temp.append('y')
				else:
					temp.append('n')
				
                
				tempp.append(temp)
				temp= []
			


					

	return tempp

import test3

def code(j, i2, u):
    # Convert j to an integer and increment by 1
    j = int(j) + 1
    
    # Assuming wb_obj is defined somewhere else in your code
    sheet = wb_obj.active
   
    # Modify cell AB{j} (assuming j is now incremented index)
    if i2 == int(float(sheet.cell(row=j, column=3).value)):
		# Assuming sheet_obj is defined somewhere else, get cell value
        sheet[f"AB{j}"] = 'استلم'
        sheet[f"AC{j}"] = f'{u}'
        majdadff = sheet_obj.cell(row=j, column=30).value
        cell_obj = sheet.cell(row=j, column=1).value
        cell_obj1 = sheet.cell(row=j, column=3).value
        cell_obj2 = sheet.cell(row=j, column=2).value
        # majdadff = sheet_obj.cell(row=d, column=30).value


        test3.create_print_image(cell_obj2, cell_obj, cell_obj1 ,majdadff)
        wb_obj.save(filename="jj.xlsx")
    else:
     
        for d in range(2, row +1): 
            cell_obj = sheet_obj.cell(row=d, column=3)
			
			
	     
            sheet[f"AB{d}"] = 'استلم'
            sheet[f"AC{d}"] = f'{u}'
            majdadf = sheet_obj.cell(row=d, column=30).value
            cell_obj = sheet.cell(row=d, column=1).value
            cell_obj1 = sheet.cell(row=d, column=3).value
            cell_obj2 = sheet.cell(row=d, column=2).value
            # majdadff = sheet_obj.cell(row=d, column=30).value
			
            test3.create_print_image(cell_obj2, cell_obj, cell_obj1, majdadf)
            wb_obj.save(filename="jj.xlsx")             				

def dcode(j, i2):
    # Convert j to an integer and increment by 1
    j = int(j) + 1
    
    # Assuming wb_obj is defined somewhere else in your code
    sheet = wb_obj.active


    # Modify cell AB{j} (assuming j is now incremented index)
    if i2 == int(float(sheet.cell(row=j, column=3).value)):
		# Assuming sheet_obj is defined somewhere else, get cell value
        sheet[f"AC{j}"] = ''
        sheet[f"AB{j}"] = ''
        # cell_obj = sheet.cell(row=j, column=1).value
        # cell_obj1 = sheet.cell(row=j, column=3).value
        # cell_obj2 = sheet.cell(row=j, column=2).value
        # test3.create_print_image(cell_obj2, cell_obj, cell_obj1)
        wb_obj.save(filename="jj.xlsx")
    else:

        for d in range(2, row +1): 
            cell_obj = sheet_obj.cell(row=d, column=3)
			
			
            sheet[f"AB{d}"] = ''        
            sheet[f"AC{d}"] = ''
            # cell_obj = sheet.cell(row=d, column=1).value
            # cell_obj1 = sheet.cell(row=d, column=3).value
            # cell_obj2 = sheet.cell(row=d, column=2).value
            # test3.create_print_image(cell_obj2, cell_obj, cell_obj1)
            wb_obj.save(filename="jj.xlsx")             				




def initcoutnt():
	tempp = []
	temp= 0
	temppp= []
	for d in range(2, row +1):
		majd = sheet_obj.cell(row=d, column=28).value
		majdd = sheet_obj.cell(row=d, column=29).value
		if majd  == 'استلم':
			temp +=1
			if majdd not in tempp:
				tempp.append(majdd)
				temppp.append(1)
			
			else:
			
				jdjd =tempp.index(majdd)
				j=temppp[jdjd]+1
				temppp[jdjd]=j
	return temp, tempp , temppp
    
