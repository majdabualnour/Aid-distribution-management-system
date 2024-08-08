import openpyxl
from openpyxl import Workbook
from datetime import datetime
def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width
input_file = 'jj.xlsx'  # Replace with your input file name


# Get the current date
current_date = datetime.now().date()

# Print the current date


output_file = f'{str(current_date)}.xlsx'  # Replace with your desired output file name
column_index = 8  # Replace with the column index you want to filter by (e.g., 2 for column B)
specific_value = 'استلم' 
def create_filtered_workbook(input_file= input_file, output_file= output_file, column_index= column_index, specific_value= specific_value):
    # Load the input workbook and sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Create a new workbook and add sheets
    new_wb = Workbook()
    sheet1 = new_wb.active
    sheet1.title = "الكل"
    source_col = 'AB'
    target_col = 8  # This corresponds to the 8th column (H)

    # Get the column indices
    source_col_index = openpyxl.utils.column_index_from_string(source_col)
    
    # Move values from source column to target column, skipping the first row
    for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
        value = sheet.cell(row=row, column=source_col_index).value
        sheet.cell(row=row, column=target_col).value = value
        sheet.cell(row=row, column=source_col_index).value = None  # Clear the original cell

    sheet2 = new_wb.create_sheet(title="استلم")
    sheet3 = new_wb.create_sheet(title="لم يستلم")
# Copy the original headers and data to sheet1
    for col in range(1, sheet.max_column + 1):
        sheet1.cell(row=1, column=col).value = sheet.cell(row=1, column=col).value

    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet1.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value


    # Copy headers to sheet2 and sheet3
    for col in range(1, sheet.max_column + 1):
        sheet2.cell(row=1, column=col + 1).value = sheet.cell(row=1, column=col).value
        sheet3.cell(row=1, column=col + 1).value = sheet.cell(row=1, column=col).value

    # Filter rows based on the specific value and add to sheet2 or sheet3
    row_filtered = 2
    row_non_filtered = 2

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value == specific_value:
            sheet2.cell(row=row_filtered, column=1).value = row_filtered - 1  # Counter column
            for col in range(1, sheet.max_column + 1):
                sheet2.cell(row=row_filtered, column=col + 1).value = sheet.cell(row=row, column=col).value
            row_filtered += 1
        else:
            sheet3.cell(row=row_non_filtered, column=1).value = row_non_filtered - 1  # Counter column
            for col in range(1, sheet.max_column + 1):
                sheet3.cell(row=row_non_filtered, column=col + 1).value = sheet.cell(row=row, column=col).value
            row_non_filtered += 1
    sheet2.delete_cols(2)
    sheet3.delete_cols(2)
    sheet2.cell(row=1, column=1).value = '#'
    sheet3.cell(row=1, column= 1).value = '#'
    # Auto-adjust column widths
    auto_adjust_column_width(sheet1)
    auto_adjust_column_width(sheet2)


    auto_adjust_column_width(sheet3)

    # Save the new workbook
    new_wb.save(output_file)
    print(f"New workbook created: {output_file}")

# Example usage
 # Replace with the specific value you want to filter by

create_filtered_workbook()
